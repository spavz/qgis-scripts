import json
import openpyxl



wb = openpyxl.load_workbook("1% Penetration Random Points.xlsx")
ws = wb.active
zones = ws["A"]
pointCounts = ws["B"]

pointCountInZone = dict()

# 400 is hardcoded arbitrary number
for i in range(400): 
    if zones[i].value is not None:
        pointCountInZone[zones[i].value] = pointCounts[i].value

# This is just to visualize dict in human readable format
# print(json.dumps(randomPointsByZoneIdDict, indent=4, sort_keys=True)) 



intersectionGeoJsonFile = "Intersection.geojson"
intersectionGeoJson = dict()
with open(intersectionGeoJsonFile) as _intersectionGeoJsonFile:
    intersectionGeoJson = json.load(_intersectionGeoJsonFile)


pointsInZone = dict()

# start iterating on features array in geojson and add them to pointsInZone, until pointsInZone reaches specified pointCount, for every zone
for point in intersectionGeoJson['features']:
    zoneId = point['properties']['Id_2']
    if pointsInZone.get(zoneId) is None:
        pointsInZone[zoneId] = []
    if len(pointsInZone[zoneId]) < pointCountInZone[zoneId]:
        pointsInZone[zoneId].append(point)


# extract feature array from every zone and make it a common all-zone feature array and dump output to json file cleanedZonePoints.json
features = []
for zoneId in pointsInZone:
    features += pointsInZone[zoneId]

with open('cleanedZonePoints.json', 'w') as cleanedZonePoints:
    json.dump({'features' : features}, cleanedZonePoints)



